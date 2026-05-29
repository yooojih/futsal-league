"""スケジュール結果をExcelファイルに出力する"""
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .multi_league import MultiScheduledMatch, MultiAssignResult
from .round_robin import Match
from .parser import AvailabilityData, date_sort_key


def _group_by_date(
    scheduled: list[MultiScheduledMatch],
) -> list[tuple[str, list[MultiScheduledMatch]]]:
    """
    MultiScheduledMatch リストを日付でグループ化し、(表示用日付文字列, 試合リスト) の
    リストを日付順で返す。

    年なし "5月3日（土）" と年あり "2026年5月3日（土）" のように同じ実日付でも
    文字列が異なる場合は同じグループにまとめる。
    表示用文字列は年情報を含む長い方を優先する。
    """
    # 各日付文字列の正規化キー (year, month, day) を計算
    date_strings: set[str] = {sm.date for sm in scheduled}
    raw: dict[str, tuple] = {d: date_sort_key(d) for d in date_strings}
    max_year = max((k[0] for k in raw.values() if k[0] > 0), default=0)

    def norm(d: str) -> tuple:
        k = raw[d]
        if k[0] == 0 and max_year > 0:
            return (max_year, k[1], k[2])
        return k

    # 同じ正規化キーを持つ日付文字列のうち最も長いものを代表表示として使う
    best_label: dict[tuple, str] = {}
    for d in date_strings:
        k = norm(d)
        if k not in best_label or len(d) > len(best_label[k]):
            best_label[k] = d

    # 試合を正規化キーでグループ化
    groups: dict[tuple, list[MultiScheduledMatch]] = defaultdict(list)
    for sm in scheduled:
        groups[norm(sm.date)].append(sm)

    return [
        (best_label[k], sorted(groups[k], key=lambda sm: sm.slot))
        for k in sorted(groups.keys())
    ]


# カラーパレット
HEADER_BG    = "2F5496"
HEADER_FG    = "FFFFFF"
SUBHEADER_BG = "D9E1F2"
MAYBE_BG     = "FFF2CC"
UNSCHEDULED_BG = "FCE4D6"
ALT_ROW_BG   = "EEF2F9"

LEAGUE_COLORS = ["4472C4", "ED7D31", "70AD47", "FFC000"]  # 青・橙・緑・黄


def _header_style(cell, color: str = HEADER_BG):
    cell.font = Font(bold=True, color=HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _leg_label(leg: int) -> str:
    return "前半戦" if leg == 1 else "後半戦"


# ------------------------------------------------------------------
# メインエントリ: 複数リーグ統合Excel
# ------------------------------------------------------------------
def build_multi_excel(
    scheduled: list[MultiScheduledMatch],
    unscheduled: dict[str, list[Match]],
    avails: dict[str, AvailabilityData],
    leagues_config: list[dict],
    period_name: str,
) -> bytes:
    wb = Workbook()

    league_names = [lc["name"] for lc in leagues_config]
    color_map = {name: LEAGUE_COLORS[i % len(LEAGUE_COLORS)] for i, name in enumerate(league_names)}

    _sheet_integrated(wb, scheduled, leagues_config, color_map, period_name)

    for lc in leagues_config:
        lname = lc["name"]
        lg_scheduled = [sm for sm in scheduled if sm.league_name == lname]
        lg_unscheduled = unscheduled.get(lname, [])
        _sheet_league(wb, lg_scheduled, lg_unscheduled, lname, color_map[lname])

    _sheet_summary(wb, scheduled, avails, league_names, color_map)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# シート①: 統合日程
# ------------------------------------------------------------------
def _sheet_integrated(
    wb: Workbook,
    scheduled: list[MultiScheduledMatch],
    leagues_config: list[dict],
    color_map: dict[str, str],
    period_name: str,
):
    ws = wb.active
    ws.title = "統合日程"

    row = 1
    ws.cell(row, 1, f"【{period_name}】統合日程").font = Font(bold=True, size=13)
    row += 2

    for date, matches in _group_by_date(scheduled):
        # 日付ヘッダー
        c = ws.cell(row, 1, f"▶ {date}（{len(matches)}試合）")
        c.font = Font(bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # 列ヘッダー
        for col, label in enumerate(["試合", "リーグ", "ホーム", "vs", "アウェイ", "前後半", "備考"], 1):
            c = ws.cell(row, col, label)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
        row += 1

        for i, sm in enumerate(matches, 1):
            lg_color = color_map.get(sm.league_name, HEADER_BG)
            bg = ALT_ROW_BG if i % 2 == 0 else "FFFFFF"
            note = "△使用" if sm.use_maybe else ""
            values = [i, sm.league_name, sm.match.home, "vs", sm.match.away, _leg_label(sm.match.leg), note]
            for col, val in enumerate(values, 1):
                c = ws.cell(row, col, val)
                c.border = _thin_border()
                c.alignment = Alignment(horizontal="center")
                if col == 2:  # リーグ列はリーグカラー
                    c.fill = PatternFill("solid", fgColor=lg_color)
                    c.font = Font(bold=True, color=HEADER_FG)
                else:
                    c.fill = PatternFill("solid", fgColor=bg)
            row += 1

        row += 1  # 日付間の空行

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8


# ------------------------------------------------------------------
# シート②: リーグ別日程
# ------------------------------------------------------------------
def _sheet_league(
    wb: Workbook,
    scheduled: list[MultiScheduledMatch],
    unscheduled: list[Match],
    league_name: str,
    color: str,
):
    ws = wb.create_sheet(f"{league_name}")

    row = 1
    ws.cell(row, 1, f"【{league_name}】スケジュール").font = Font(bold=True, size=12)
    row += 2

    for date, matches in _group_by_date(scheduled):
        c = ws.cell(row, 1, f"▶ {date}")
        c.font = Font(bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        for col, label in enumerate(["スロット", "ホーム", "vs", "アウェイ", "前後半", "備考"], 1):
            c = ws.cell(row, col, label)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
        row += 1

        for sm in matches:
            bg = ALT_ROW_BG if sm.slot % 2 == 0 else "FFFFFF"
            note = "△使用" if sm.use_maybe else ""
            values = [f"第{sm.slot}試合", sm.match.home, "vs", sm.match.away, _leg_label(sm.match.leg), note]
            for col, val in enumerate(values, 1):
                c = ws.cell(row, col, val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center")
                c.border = _thin_border()
            row += 1

        row += 1

    if unscheduled:
        ws.cell(row, 1, "【未割り当て試合】").font = Font(bold=True, color="CC0000")
        row += 1
        for m in unscheduled:
            c = ws.cell(row, 1, f"{m.home} vs {m.away}（{_leg_label(m.leg)}）")
            c.fill = PatternFill("solid", fgColor=UNSCHEDULED_BG)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1

    for col, width in zip("ABCDEF", [10, 26, 5, 26, 8, 8]):
        ws.column_dimensions[col].width = width


# ------------------------------------------------------------------
# シート③: 消化数サマリ（全リーグ）
# ------------------------------------------------------------------
def _sheet_summary(
    wb: Workbook,
    scheduled: list[MultiScheduledMatch],
    avails: dict[str, AvailabilityData],
    league_names: list[str],
    color_map: dict[str, str],
):
    ws = wb.create_sheet("消化数サマリ")

    # リーグごとに消化数を集計
    gc: dict[str, dict[str, int]] = {name: {} for name in league_names}
    for sm in scheduled:
        d = gc[sm.league_name]
        d[sm.match.home] = d.get(sm.match.home, 0) + 1
        d[sm.match.away] = d.get(sm.match.away, 0) + 1

    row = 1
    for lname in league_names:
        color = color_map[lname]
        c = ws.cell(row, 1, lname)
        _header_style(c, color)
        ws.cell(row, 2, "消化試合数").font = Font(bold=True)
        _header_style(ws.cell(row, 2), color)
        row += 1

        counts = gc[lname]
        teams = avails[lname].teams if lname in avails else list(counts.keys())
        max_count = max(counts.values(), default=0)

        for team in teams:
            count = counts.get(team, 0)
            ws.cell(row, 1, team).border = _thin_border()
            c = ws.cell(row, 2, count)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            if count == max_count and max_count > 0:
                c.fill = PatternFill("solid", fgColor="C6EFCE")
            row += 1

        row += 1  # リーグ間の空行

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
